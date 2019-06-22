# -*- coding: utf-8 -*-

import csv
import io
import sys
import docx
from openpyxl import load_workbook
from pyloco import Task

class Xlsx2Docxtable(Task):
    """converts Microsoft Excel file to a MS Word file.

'xlsx2docxtable' task a sheet of Microsoft Excel file to a table in a MS Word file.

Example(s)
----------

Current version of the task assumes that an input Excel file is generated
by 'docxtable2xlsx' from an input Word file.

Follwoing command reads tables.xlsx Excel file and my.docx MS word file,
and convert Excel sheets in the file tables.xlsx to tables of MS Word file of out.docx. ::

    >>> pyloco xlsx2docxtable tables.xlsx my.docx -o out.docx
    out.docx 

Follwoing command reads tables.csv CSV file instead of Excel file in above example. ::

    >>> pyloco xlsx2docxtable tables.csv my.docx -o out.docx
    out.docx 
"""

    _name_ = "xlsx2docxtable"
    _version_ = "0.1.0"
    _install_requires_ = ["python-docx", "openpyxl"]

    def __init__(self, parent):

        self.add_data_argument("xlsx", type=str, help="input xlsx file")
        self.add_data_argument("docx", type=str, help="input docx file")

        self.add_option_argument("-t", "--type", metavar="type",
                default="xlsx", 
                help="input file format (default='xlsx')") 

        self.add_option_argument(
            "-o", "--output", default="out.docx", help=("output file")
        )

        self.register_forward("data", help="output data")

    def perform(self, targs):

        tables = {}

        if targs.type == "xlsx":
            wb = load_workbook(targs.xlsx)
            for sidx, ws in enumerate(wb.worksheets):
                if sidx in tables:
                    table = tables[sidx]

                else:
                    table = {}
                    tables[sidx] = table

                for ridx, row in enumerate(ws.iter_rows()):
                    if ridx in table:
                        r = table[ridx]

                    else:
                        r = {}
                        table[ridx] = r

                    for cidx, col in enumerate(row):
                        r[cidx] = col.value

        elif targs.type == "csv":

            with io.open(targs.xlsx, 'r', encoding="utf-8") as csvfile:
                reader = csv.reader(csvfile, delimiter=',',
                                        quotechar='|', quoting=csv.QUOTE_MINIMAL)
                header = reader.__next__()

                for row in reader:
                    tid, rid, cid = tuple(int(i) for i in row[:-1])
                    val = row[-1]

                    if tid in tables:
                        table = tables[tid]

                    else:
                        table = {}
                        tables[tid] = table


                    if rid in table:
                        row = table[rid]

                    else:
                        row = {}
                        table[rid] = row

                    row[cid] = val

        else:
            print("Unknown output type: %s" % targs.type)
            sys.exit(1)

        docf = docx.Document(targs.docx)

        for tidx, table in enumerate(docf.tables):
            for ridx, row in enumerate(table.rows):
                for cidx, cell in enumerate(row.cells):
                    if tidx not in tables:
                        continue
                    if ridx not in tables[tidx]:
                        continue
                    if cidx not in tables[tidx][ridx]:
                        continue
                    cell.text = tables[tidx][ridx][cidx]

        docf.save(targs.output)
        self.add_forward(data=docf)
